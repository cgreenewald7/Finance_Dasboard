import matplotlib.pyplot as plt
import matplotlib.backends.backend_tkagg as tkagg
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import Calendar
from datetime import datetime, timedelta
import csv
import os
from pathlib import Path

class BudgetTracker:
    def __init__(self):
        # user specific directory for CSV file
        home_dir = Path.home()
        data_dir = home_dir / "Documents" / "BudgetTracker"
        data_dir.mkdir(parents=True, exist_ok=True)
        self.data_file = data_dir / "budget_data.csv"            

        # self.data_file = "budget_data.csv"
        self.default_categories = [
            "Activities", "Bills/Utilities", "Charitable", "Eating Out",
            "Groceries", "Housing", "Other", "Savings/Investing",
            "Shopping", "Transportation"
        ]
        self.categories = list(self.default_categories)
        self.payment_methods = ["Debit Card", "Credit Card", "Cash", "Bank Transfer", "Other"]
        self.saving_categories = ["Emergency Fund", "Retirement", "Brokerage", "Education", "Vacation Fund", "Other"]
        self.category_goals = {}
        self.category_goal_types = {}
        self.deleted_categories = set()
        self.dark_mode = False
        self.colors = self.get_palette(self.dark_mode)
        
        self.root = tk.Tk()
        self.root.title("Budget Tracker")
        self.root.geometry("1120x860")
        self.root.minsize(980, 760)
        self.root.configure(bg=self.colors["bg"])
        
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.configure_styles()

        self.showing_list = False
        self.current_month = datetime.now().strftime("%Y-%m")
        self.budget_canvas = None
        self.all_income = []  # Store all income across months
        self.all_expenses = []  # Store all expenses across months
        self.all_savings = []  # Store saving and investing transfers across months
        
        self.load_data()
        self.setup_gui()

    def get_palette(self, dark_mode):
        if dark_mode:
            return {
                "bg": "#111827",
                "surface": "#1f2937",
                "surface_alt": "#374151",
                "text": "#f9fafb",
                "muted": "#cbd5e1",
                "primary": "#2f9faf",
                "primary_dark": "#237b88",
                "accent": "#f2a65a",
                "good": "#5cc98b",
                "bad": "#f87171",
                "border": "#4b5563",
                "chart": "#1f2937",
            }
        return {
            "bg": "#f4f7fb",
            "surface": "#ffffff",
            "surface_alt": "#eef3f8",
            "text": "#17202a",
            "muted": "#5d6d7e",
            "primary": "#1f7a8c",
            "primary_dark": "#155e6d",
            "accent": "#bf6f21",
            "good": "#247a4d",
            "bad": "#b33a3a",
            "border": "#d8e1ea",
            "chart": "#ffffff",
        }

    def configure_styles(self):
        self.style.configure("TButton", 
                            font=("Segoe UI", 11, "bold"),
                            background=self.colors["primary"],
                            foreground="white",
                            borderwidth=0,
                            focusthickness=0,
                            padding=(14, 9))
        self.style.map("TButton",
                      background=[('active', self.colors["primary_dark"])])
        
        # Modernize Combobox styling
        self.style.configure("TCombobox",
                            fieldbackground=self.colors["surface"],
                            background=self.colors["surface"],
                            foreground=self.colors["text"],
                            arrowcolor="white",
                            borderwidth=0,
                            font=("Segoe UI", 11),
                            padding=5)
        self.style.map("TCombobox",
                    fieldbackground=[('readonly', self.colors["surface"])],
                    background=[('readonly', self.colors["surface"]), ('active', self.colors["surface_alt"])],
                    foreground=[('readonly', self.colors["text"])])
        self.style.configure("TEntry",
                            fieldbackground=self.colors["surface"],
                            foreground=self.colors["text"],
                            borderwidth=1,
                            padding=6)
        
        self.style.configure("Treeview",
                            background=self.colors["surface"],
                            foreground=self.colors["text"],
                            fieldbackground=self.colors["surface"],
                            font=("Segoe UI", 10),
                            rowheight=30,
                            borderwidth=0,
                            relief="flat",
                            highlightthickness=0)
        self.style.configure("Treeview.Heading",
                            font=("Segoe UI", 10, "bold"),
                            background=self.colors["surface_alt"],
                            foreground=self.colors["text"],
                            borderwidth=1,
                            relief="solid")
        self.style.map("Treeview",
                      background=[('selected', self.colors["primary"])],
                      foreground=[('selected', 'white')])
        self.style.layout("Treeview", [
            ('Treeview.treearea', {'sticky': 'nswe'})
        ])

    def toggle_theme(self):
        self.dark_mode = not self.dark_mode
        self.colors = self.get_palette(self.dark_mode)
        self.root.configure(bg=self.colors["bg"])
        self.configure_styles()
        for widget in self.root.winfo_children():
            widget.destroy()
        self.setup_gui()

    def add_category(self, category):
        category = category.strip() or "Other"
        category = self.normalize_category(category)
        existing = {item.lower(): item for item in self.categories}
        if category.lower() not in existing:
            self.categories.append(category)
            self.categories.sort(key=str.lower)
        self.deleted_categories.discard(category)
        return existing.get(category.lower(), category)

    def normalize_category(self, category):
        if category.strip().lower() in ("saving", "savings", "investing", "investment", "investments"):
            return "Savings/Investing"
        return category.strip() or "Other"

    def get_expense_category_options(self, current_category=None):
        options = [category for category in self.categories if category != "Savings/Investing"]
        if current_category and current_category not in options:
            options.append(current_category)
        return sorted(options, key=str.lower)

    def chart_colors(self, count):
        if self.dark_mode:
            palette = (
                "#6aa8d8", "#58b7ba", "#7fc7a5", "#8f9fe0", "#54a9c3",
                "#9ab0c8", "#7db3d2", "#66bda9", "#93a9df", "#76c0c2",
                "#a7bad0", "#7299c8"
            )
        else:
            palette = (
                "#4f7cac", "#4b9aaa", "#6f9f8f", "#6c83b5", "#3f8f9f",
                "#7a90a8", "#5f8fb4", "#5d9b8f", "#7896c7", "#6fa8aa",
                "#8aa0b8", "#557ca3"
            )
        return [palette[i % len(palette)] for i in range(count)]

    def center_text_style(self):
        return {
            "ha": "center",
            "va": "center",
            "fontsize": 16,
            "fontweight": "bold",
            "color": self.colors["text"],
            "bbox": {
                "boxstyle": "round,pad=0.45",
                "facecolor": self.colors["chart"],
                "edgecolor": self.colors["border"],
                "linewidth": 1,
                "alpha": 0.92,
            },
        }

    def load_data(self):
        self.categories = list(self.default_categories)
        self.category_goals = {}
        self.category_goal_types = {}
        self.deleted_categories = set()
        self.all_income = []
        self.all_expenses = []
        self.all_savings = []
        category_config_loaded = False
        if os.path.exists(self.data_file):
            with open(self.data_file, mode='r', newline='') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    if row["type"] == "income":
                        self.all_income.append({
                            "source": row["source"],
                            "amount": float(row["amount"]),
                            "date": row["date"],
                            "month": row["date"][:7]
                        })
                    elif row["type"] == "expense":
                        category = self.normalize_category(row["category"])
                        self.all_expenses.append({
                            "where": row["where"],
                            "amount": float(row["amount"]),
                            "date": row["date"],
                            "category": category,
                            "payment_method": row.get("payment_method", "Debit Card") or "Debit Card",
                            "authorized_user": row.get("authorized_user", "") or "",
                            "month": row["date"][:7]  # Extract YYYY-MM
                        })
                    elif row["type"] in ("saving", "savings", "saving_investment"):
                        saving_category = row.get("category", "Other") or "Other"
                        self.all_savings.append({
                            "account": row.get("where", "") or row.get("source", ""),
                            "amount": float(row["amount"]),
                            "date": row["date"],
                            "category": saving_category,
                            "month": row["date"][:7]
                        })
                        if saving_category not in self.saving_categories:
                            self.saving_categories.append(saving_category)
                            self.saving_categories.sort(key=str.lower)
                    elif row["type"] == "category_config":
                        if not category_config_loaded:
                            self.categories = []
                            self.category_goals = {}
                            self.category_goal_types = {}
                            category_config_loaded = True
                        category = self.normalize_category(row.get("category", ""))
                        if category:
                            self.add_category(category)
                            goal = row.get("goal_amount", "")
                            if goal:
                                self.category_goals[category] = float(goal)
                                goal_type = row.get("goal_type", "") or "amount"
                                self.category_goal_types[category] = goal_type if goal_type in ("amount", "percent") else "amount"
                    elif row["type"] == "category_deleted":
                        raw_category = row.get("category", "").strip()
                        if raw_category.lower() in ("saving", "savings", "investing", "investment", "investments"):
                            continue
                        category = self.normalize_category(raw_category)
                        if category:
                            self.deleted_categories.add(category)

        for category in self.default_categories:
            if category not in self.categories and category not in self.deleted_categories:
                self.add_category(category)
        self.categories = sorted(self.categories, key=str.lower)

    def save_data(self):
        with open(self.data_file, mode='w', newline='') as file:
            fieldnames = [
                "type", "source", "where", "amount", "date", "category",
                "payment_method", "authorized_user", "goal_group", "goal_amount", "goal_type"
            ]
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            for inc in self.all_income:
                writer.writerow({
                    "type": "income",
                    "source": inc["source"],
                    "where": "",
                    "amount": inc["amount"],
                    "date": inc["date"],
                    "category": "",
                    "payment_method": "",
                    "authorized_user": "",
                    "goal_group": "",
                    "goal_amount": "",
                    "goal_type": ""
                })
            for exp in self.all_expenses:
                writer.writerow({
                    "type": "expense",
                    "source": "",
                    "where": exp["where"],
                    "amount": exp["amount"],
                    "date": exp["date"],
                    "category": exp["category"],
                    "payment_method": exp.get("payment_method", "Debit Card"),
                    "authorized_user": exp.get("authorized_user", ""),
                    "goal_group": "",
                    "goal_amount": "",
                    "goal_type": ""
                })
            for item in self.all_savings:
                writer.writerow({
                    "type": "saving_investment",
                    "source": "",
                    "where": item["account"],
                    "amount": item["amount"],
                    "date": item["date"],
                    "category": item["category"],
                    "payment_method": "",
                    "authorized_user": "",
                    "goal_group": "",
                    "goal_amount": "",
                    "goal_type": ""
                })
            for category in self.categories:
                writer.writerow({
                    "type": "category_config",
                    "source": "",
                    "where": "",
                    "amount": "",
                    "date": "",
                    "category": category,
                    "payment_method": "",
                    "authorized_user": "",
                    "goal_group": "",
                    "goal_amount": self.category_goals.get(category, ""),
                    "goal_type": self.category_goal_types.get(category, "amount") if category in self.category_goals else ""
                })
            for category in sorted(self.deleted_categories, key=str.lower):
                writer.writerow({
                    "type": "category_deleted",
                    "source": "",
                    "where": "",
                    "amount": "",
                    "date": "",
                    "category": category,
                    "payment_method": "",
                    "authorized_user": "",
                    "goal_group": "",
                    "goal_amount": "",
                    "goal_type": ""
                })

    def get_month_data(self, month):
        """Filter income and expenses for the given month."""
        income = [i for i in self.all_income if i["month"] == month]
        expenses = [e for e in self.all_expenses if e["month"] == month]
        return income, expenses

    def get_month_savings(self, month):
        return [s for s in self.all_savings if s["month"] == month]

    def get_savings_investing_total(self, expenses, savings):
        expense_savings = sum(
            exp["amount"] for exp in expenses
            if self.normalize_category(exp.get("category", "")) == "Savings/Investing"
        )
        transfer_savings = sum(item["amount"] for item in savings)
        return expense_savings + transfer_savings

    def get_expense_breakdown_totals(self, expenses, savings=None):
        category_totals = {}
        for exp in expenses:
            cat = self.normalize_category(exp["category"])
            category_totals[cat] = category_totals.get(cat, 0) + exp["amount"]
        return category_totals

    def export_selected_month(self):
        income, expenses = self.get_month_data(self.current_month)
        savings = self.get_month_savings(self.current_month)
        if not income and not expenses and not savings:
            messagebox.showinfo("Export Month", f"No data found for {self.current_month}.")
            return

        default_name = f"budget_{self.current_month}.csv"
        file_path = filedialog.asksaveasfilename(
            title="Export selected month",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not file_path:
            return

        fieldnames = [
            "Type", "Date", "Month", "Source/Recipient", "Amount",
            "Category", "Payment Method", "Authorized User"
        ]
        with open(file_path, mode="w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            for inc in sorted(income, key=lambda x: x["date"]):
                writer.writerow({
                    "Type": "Income",
                    "Date": inc["date"],
                    "Month": self.current_month,
                    "Source/Recipient": inc["source"],
                    "Amount": f"{inc['amount']:.2f}",
                    "Category": "",
                    "Payment Method": "",
                    "Authorized User": ""
                })
            for exp in sorted(expenses, key=lambda x: x["date"]):
                writer.writerow({
                    "Type": "Expense",
                    "Date": exp["date"],
                    "Month": self.current_month,
                    "Source/Recipient": exp["where"],
                    "Amount": f"{exp['amount']:.2f}",
                    "Category": exp["category"],
                    "Payment Method": exp.get("payment_method", "Debit Card"),
                    "Authorized User": exp.get("authorized_user", "")
                })
            for item in sorted(savings, key=lambda x: x["date"]):
                writer.writerow({
                    "Type": "Saving/Investment",
                    "Date": item["date"],
                    "Month": self.current_month,
                    "Source/Recipient": item["account"],
                    "Amount": f"{item['amount']:.2f}",
                    "Category": item["category"],
                    "Payment Method": "",
                    "Authorized User": ""
                })

        messagebox.showinfo("Export Complete", f"Exported {self.current_month} to:\n{file_path}")

    def export_selected_month_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Excel Export Unavailable",
                "The openpyxl package is required for Excel export. Install it with: pip install openpyxl"
            )
            return

        income, expenses = self.get_month_data(self.current_month)
        savings = self.get_month_savings(self.current_month)
        if not income and not expenses and not savings:
            messagebox.showinfo("Export Month", f"No data found for {self.current_month}.")
            return

        default_name = f"budget_{self.current_month}.xlsx"
        file_path = filedialog.asksaveasfilename(
            title="Export selected month to Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")]
        )
        if not file_path:
            return

        wb = Workbook()
        transactions_ws = wb.active
        transactions_ws.title = "Transactions"
        headers = [
            "Type", "Date", "Month", "Source/Recipient", "Amount",
            "Category", "Payment Method", "Authorized User"
        ]
        transactions_ws.append(headers)
        for cell in transactions_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F7CAC")

        for inc in sorted(income, key=lambda x: x["date"]):
            transactions_ws.append([
                "Income", inc["date"], self.current_month, inc["source"],
                inc["amount"], "", "", ""
            ])
        for exp in sorted(expenses, key=lambda x: x["date"]):
            transactions_ws.append([
                "Expense", exp["date"], self.current_month, exp["where"],
                exp["amount"], exp["category"],
                exp.get("payment_method", "Debit Card"), exp.get("authorized_user", "")
            ])
        for item in sorted(savings, key=lambda x: x["date"]):
            transactions_ws.append([
                "Saving/Investment", item["date"], self.current_month, item["account"],
                item["amount"], item["category"], "", ""
            ])

        for col in range(1, len(headers) + 1):
            transactions_ws.column_dimensions[get_column_letter(col)].width = 18
        for row in transactions_ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = '$#,##0.00'

        expense_totals = self.get_expense_breakdown_totals(expenses, savings)

        total_income = sum(inc["amount"] for inc in income)
        total_expenses = sum(exp["amount"] for exp in expenses)
        total_saved = self.get_savings_investing_total(expenses, savings)
        net_total = total_income - total_expenses

        summary_start = transactions_ws.max_row + 3
        transactions_ws.cell(row=summary_start, column=1, value="Monthly Totals")
        transactions_ws.cell(row=summary_start, column=1).font = Font(bold=True, size=14)
        totals_headers = ["Income", "Expenses", "Net", "Saved / Invested"]
        totals_values = [total_income, total_expenses, net_total, total_saved]
        for col, header in enumerate(totals_headers, start=1):
            cell = transactions_ws.cell(row=summary_start + 1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F7CAC")
        for col, value in enumerate(totals_values, start=1):
            cell = transactions_ws.cell(row=summary_start + 2, column=col, value=value)
            cell.number_format = '$#,##0.00'
            cell.font = Font(bold=True)

        category_start = summary_start + 5
        transactions_ws.cell(row=category_start, column=1, value="Expense Category Summary")
        transactions_ws.cell(row=category_start, column=1).font = Font(bold=True, size=14)
        transactions_ws.cell(row=category_start + 1, column=1, value="Category")
        transactions_ws.cell(row=category_start + 1, column=2, value="Total Spending")
        for cell in transactions_ws[category_start + 1][:2]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F7CAC")

        row_index = category_start + 2
        for category, amount in sorted(expense_totals.items(), key=lambda item: item[0].lower()):
            transactions_ws.cell(row=row_index, column=1, value=category)
            transactions_ws.cell(row=row_index, column=2, value=amount)
            transactions_ws.cell(row=row_index, column=2).number_format = '$#,##0.00'
            row_index += 1

        transactions_ws.cell(row=row_index, column=1, value="Grand Total")
        transactions_ws.cell(row=row_index, column=2, value=sum(expense_totals.values()))
        transactions_ws.cell(row=row_index, column=1).font = Font(bold=True)
        transactions_ws.cell(row=row_index, column=2).font = Font(bold=True)
        transactions_ws.cell(row=row_index, column=2).number_format = '$#,##0.00'

        wb.save(file_path)
        messagebox.showinfo("Export Complete", f"Exported Excel workbook to:\n{file_path}")

    def setup_gui(self):
        header_frame = tk.Frame(self.root, bg=self.colors["surface"], padx=24, pady=16,
                                highlightbackground=self.colors["border"], highlightthickness=1)
        header_frame.pack(fill="x")
        
        title_label = tk.Label(header_frame, 
                            text="Budget Tracker",
                            font=("Segoe UI", 24, "bold"),
                            fg=self.colors["text"],
                            bg=self.colors["surface"])
        title_label.pack(side="left")

        subtitle_label = tk.Label(header_frame,
                                text=f"Local CSV: {self.data_file}",
                                font=("Segoe UI", 9),
                                fg=self.colors["muted"],
                                bg=self.colors["surface"])
        subtitle_label.pack(side="left", padx=(12, 0), pady=(8, 0))

        self.month_var = tk.StringVar(value=self.current_month)
        self.month_dropdown = ttk.Combobox(header_frame, textvariable=self.month_var, 
                                        values=self.get_month_options(), state="readonly", width=12)
        self.month_dropdown.pack(side="right", padx=(10, 0))
        self.month_dropdown.bind("<<ComboboxSelected>>", self.update_month_view)
        tk.Label(header_frame, text="Month", font=("Segoe UI", 10, "bold"),
                fg=self.colors["muted"], bg=self.colors["surface"]).pack(side="right")
        
        button_frame = tk.Frame(self.root, bg=self.colors["bg"], padx=24, pady=14)
        button_frame.pack(fill="x")
        
        ttk.Button(button_frame, text="Add Income", 
                command=self.open_income_window, style="TButton").pack(side="left", padx=(0, 10))
        ttk.Button(button_frame, text="Add Expense", 
                command=self.open_expense_window, style="TButton").pack(side="left", padx=(0, 10))
        ttk.Button(button_frame, text="Add Saving/Investment",
                command=self.open_saving_window, style="TButton").pack(side="left", padx=(0, 10))
        ttk.Button(button_frame, text="Goals & Categories",
                command=self.open_goals_window, style="TButton").pack(side="left", padx=(0, 10))
        ttk.Button(button_frame, text="Analyze", 
                command=self.analyze_months, style="TButton").pack(side="left", padx=(0, 10))
        ttk.Button(button_frame, text="Export Month CSV",
                command=self.export_selected_month, style="TButton").pack(side="left", padx=(0, 10))
        ttk.Button(button_frame, text="Export Month Excel",
                command=self.export_selected_month_excel, style="TButton").pack(side="left", padx=(0, 10))
        self.theme_button = ttk.Button(button_frame,
                text="Dark Mode" if not self.dark_mode else "Light Mode",
                command=self.toggle_theme, style="TButton")
        self.theme_button.pack(side="right")

        self.summary_frame = tk.Frame(self.root, bg=self.colors["bg"], padx=24)
        self.summary_frame.pack(fill="x", pady=(0, 8))
        self.summary_labels = {}
        self.summary_cards = {}
        for key, label in (
            ("income", "Income"),
            ("expenses", "Expenses"),
            ("net", "Net"),
            ("savings", "Saved / Invested")
        ):
            card = tk.Frame(self.summary_frame, bg=self.colors["surface"], padx=18, pady=12,
                            highlightbackground=self.colors["border"], highlightthickness=1)
            card.pack(side="left", fill="x", expand=True, padx=(0, 10))
            tk.Label(card, text=label, font=("Segoe UI", 9, "bold"), fg=self.colors["muted"],
                    bg=self.colors["surface"]).pack(anchor="w")
            value = tk.Label(card, text="$0.00", font=("Segoe UI", 18, "bold"),
                            fg=self.colors["text"], bg=self.colors["surface"])
            value.pack(anchor="w", pady=(3, 0))
            self.summary_labels[key] = value
            self.summary_cards[key] = card
        
        self.chart_frame = tk.Frame(self.root, bg=self.colors["bg"])
        self.chart_frame.pack(fill="both", expand=True, padx=24, pady=10)
        
        self.chart_frame.grid_columnconfigure((0, 1), weight=1)
        
        # Income Pie Chart
        self.income_fig, self.income_ax = plt.subplots(figsize=(5, 5))
        self.income_fig.patch.set_facecolor(self.colors["bg"])
        self.income_ax.set_facecolor(self.colors["chart"])
        self.income_canvas = tkagg.FigureCanvasTkAgg(self.income_fig, master=self.chart_frame)
        self.income_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew", padx=10)
        self.income_fig.canvas.mpl_connect('motion_notify_event', self.on_income_hover)
        self.income_fig.canvas.mpl_connect('button_press_event', self.on_income_click)

        # Expense Pie Chart
        self.expense_fig, self.expense_ax = plt.subplots(figsize=(5, 5))
        self.expense_fig.patch.set_facecolor(self.colors["bg"])
        self.expense_ax.set_facecolor(self.colors["chart"])
        self.expense_canvas = tkagg.FigureCanvasTkAgg(self.expense_fig, master=self.chart_frame)
        self.expense_canvas.get_tk_widget().grid(row=0, column=1, sticky="nsew", padx=10)
        self.expense_fig.canvas.mpl_connect('motion_notify_event', self.on_expense_hover)
        self.expense_fig.canvas.mpl_connect('button_press_event', self.on_expense_click)
        
        # View Buttons
        self.view_income_button = ttk.Button(self.chart_frame, text="View Income List", 
                                            command=self.toggle_income_view, style="TButton")
        self.view_income_button.grid(row=1, column=0, pady=5)
        
        self.view_expense_button = ttk.Button(self.chart_frame, text="View Expense List", 
                                            command=self.toggle_expense_view, style="TButton")
        self.view_expense_button.grid(row=1, column=1, pady=5)

        self.view_savings_button = ttk.Button(self.chart_frame, text="View Saved/Invested List",
                                            command=self.toggle_savings_view, style="TButton")
        self.view_savings_button.grid(row=2, column=0, columnspan=2, pady=(0, 8))
        
        # Income Table Frame
        income_table_frame = tk.Frame(self.chart_frame, bg=self.colors["surface"], padx=8, pady=8,
                                      highlightbackground=self.colors["border"], highlightthickness=1)
        income_table_frame.grid(row=0, column=0, sticky="nsew", padx=10)
        income_table_frame.grid_remove()
        
        self.income_tree = ttk.Treeview(income_table_frame, 
                                    columns=("Date", "Source", "Amount"), 
                                    show="headings", height=20)
        self.income_tree.heading("Date", text="Date")
        self.income_tree.heading("Source", text="Source")
        self.income_tree.heading("Amount", text="Amount")
        uniform_width = 100
        self.income_tree.column("Date", width=uniform_width, anchor="center")
        self.income_tree.column("Source", width=uniform_width, anchor="w")
        self.income_tree.column("Amount", width=uniform_width, anchor="e")
        self.income_tree.pack(side="top", fill="both", expand=True)
        
        income_scrollbar = ttk.Scrollbar(income_table_frame, orient="vertical", command=self.income_tree.yview)
        income_scrollbar.pack(side="right", fill="y")
        self.income_tree.configure(yscrollcommand=income_scrollbar.set)
        
        income_button_frame = tk.Frame(income_table_frame, bg=self.colors["surface"])
        income_button_frame.pack(side="bottom", pady=5)
        ttk.Button(income_button_frame, text="Edit", command=self.edit_income, style="TButton").pack(side="left", padx=5)
        ttk.Button(income_button_frame, text="Delete", command=self.delete_income, style="TButton").pack(side="left", padx=5)
        
        self.income_table_frame = income_table_frame

        # Expense Table Frame
        expense_table_frame = tk.Frame(self.chart_frame, bg=self.colors["surface"], padx=8, pady=8,
                                       highlightbackground=self.colors["border"], highlightthickness=1)
        expense_table_frame.grid(row=0, column=1, sticky="nsew", padx=10)
        expense_table_frame.grid_remove()
        
        self.expense_tree = ttk.Treeview(expense_table_frame, 
                                        columns=("Date", "Recipient", "Amount", "Category", "Payment", "Authorized User"), 
                                        show="headings", height=20)
        self.expense_tree.heading("Date", text="Date")
        self.expense_tree.heading("Recipient", text="Recipient")
        self.expense_tree.heading("Amount", text="Amount")
        self.expense_tree.heading("Category", text="Category")
        self.expense_tree.heading("Payment", text="Payment")
        self.expense_tree.heading("Authorized User", text="Authorized User")
        self.expense_tree.column("Date", width=86, anchor="center")
        self.expense_tree.column("Recipient", width=120, anchor="w")
        self.expense_tree.column("Amount", width=82, anchor="e")
        self.expense_tree.column("Category", width=110, anchor="w")
        self.expense_tree.column("Payment", width=105, anchor="w")
        self.expense_tree.column("Authorized User", width=125, anchor="w")
        self.expense_tree.pack(side="top", fill="both", expand=True)
        
        expense_scrollbar = ttk.Scrollbar(expense_table_frame, orient="vertical", command=self.expense_tree.yview)
        expense_scrollbar.pack(side="right", fill="y")
        self.expense_tree.configure(yscrollcommand=expense_scrollbar.set)
        
        expense_button_frame = tk.Frame(expense_table_frame, bg=self.colors["surface"])
        expense_button_frame.pack(side="bottom", pady=5)
        ttk.Button(expense_button_frame, text="Edit", command=self.edit_expense, style="TButton").pack(side="left", padx=5)
        ttk.Button(expense_button_frame, text="Delete", command=self.delete_expense, style="TButton").pack(side="left", padx=5)
        
        self.expense_table_frame = expense_table_frame

        # Saved / Invested Table Frame
        savings_table_frame = tk.Frame(self.chart_frame, bg=self.colors["surface"], padx=8, pady=8,
                                       highlightbackground=self.colors["border"], highlightthickness=1)
        savings_table_frame.grid(row=0, column=0, columnspan=2, pady=(34, 0))
        savings_table_frame.grid_remove()
        savings_table_frame.configure(width=580, height=330)
        savings_table_frame.pack_propagate(False)

        savings_tree_frame = tk.Frame(savings_table_frame, bg=self.colors["surface"])
        savings_tree_frame.pack(side="top", fill="both", expand=True)
        self.savings_tree = ttk.Treeview(savings_tree_frame,
                                        columns=("Date", "Account", "Amount", "Type"),
                                        show="headings", height=8)
        self.savings_tree.heading("Date", text="Date")
        self.savings_tree.heading("Account", text="Account")
        self.savings_tree.heading("Amount", text="Amount")
        self.savings_tree.heading("Type", text="Type")
        self.savings_tree.column("Date", width=100, anchor="center")
        self.savings_tree.column("Account", width=190, anchor="w")
        self.savings_tree.column("Amount", width=100, anchor="e")
        self.savings_tree.column("Type", width=150, anchor="w")
        self.savings_tree.pack(side="left", fill="both", expand=True)

        savings_scrollbar = ttk.Scrollbar(savings_tree_frame, orient="vertical", command=self.savings_tree.yview)
        savings_scrollbar.pack(side="right", fill="y")
        self.savings_tree.configure(yscrollcommand=savings_scrollbar.set)

        savings_button_frame = tk.Frame(savings_table_frame, bg=self.colors["surface"])
        savings_button_frame.pack(side="bottom", pady=5)
        ttk.Button(savings_button_frame, text="Edit", command=self.edit_saving, style="TButton").pack(side="left", padx=5)
        ttk.Button(savings_button_frame, text="Delete", command=self.delete_saving, style="TButton").pack(side="left", padx=5)
        ttk.Button(savings_button_frame, text="Close", command=self.close_savings_view, style="TButton").pack(side="left", padx=5)

        self.savings_table_frame = savings_table_frame

        # Analysis Table Frame
        analysis_table_frame = tk.Frame(self.chart_frame, bg=self.colors["surface"], padx=8, pady=8,
                                        highlightbackground=self.colors["border"], highlightthickness=1)
        analysis_table_frame.grid(row=0, column=0, sticky="nsew", padx=10)
        analysis_table_frame.grid_remove()
        
        self.analysis_tree = ttk.Treeview(analysis_table_frame, 
                                        columns=("Analysis"), 
                                        show="tree", height=20)  # Single column, no headings
        self.analysis_tree.column("Analysis", width=300, anchor="w")
        self.analysis_tree.pack(side="top", fill="both", expand=True)
        
        analysis_scrollbar = ttk.Scrollbar(analysis_table_frame, orient="vertical", command=self.analysis_tree.yview)
        analysis_scrollbar.pack(side="right", fill="y")
        self.analysis_tree.configure(yscrollcommand=analysis_scrollbar.set)
        
        analysis_button_frame = tk.Frame(analysis_table_frame, bg=self.colors["surface"])
        analysis_button_frame.pack(side="bottom", pady=5)
        ttk.Button(analysis_button_frame, text="Close Analysis", 
                command=self.close_analysis, style="TButton").pack(pady=5)
        
        self.analysis_table_frame = analysis_table_frame

        self.budget_frame = tk.Frame(self.root, bg=self.colors["bg"])
        self.budget_frame.pack(fill="x", pady=5)

        self.net_frame = tk.Frame(self.root, bg=self.colors["bg"])
        self.net_frame.pack(fill="x", pady=5)
        self.net_label = tk.Label(self.net_frame, text="Net Income: $0.00", 
                                font=("Segoe UI", 14, "bold"), bg=self.colors["bg"])
        self.net_label.pack()

        self.update_charts()

    def get_month_options(self):
        # Include all months from data plus current month
        months = (
            set(i["month"] for i in self.all_income)
            | set(e["month"] for e in self.all_expenses)
            | set(s["month"] for s in self.all_savings)
        )
        current_date = datetime.now()
        for i in range(-12, 7):
            month_date = current_date + timedelta(days=30 * i)
            months.add(month_date.strftime("%Y-%m"))
        return sorted(months, reverse=True)

    def update_month_view(self, event):
        self.current_month = self.month_var.get()
        self.update_charts()

    def open_income_window(self):
        window = tk.Toplevel(self.root)
        window.title("Add Income")
        window.geometry("400x350")
        window.configure(bg=self.colors["surface"])
        
        tk.Label(window, text="New Income", font=("Arial", 16, "bold"), 
                fg=self.colors["text"], bg=self.colors["surface"]).pack(pady=10)
        
        frame = tk.Frame(window, bg=self.colors["surface"])
        frame.pack(pady=10)
        
        tk.Label(frame, text="Source:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=0, column=0, padx=5, pady=5)
        source_entry = ttk.Entry(frame)
        source_entry.grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(frame, text="Amount:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=1, column=0, padx=5, pady=5)
        amount_entry = ttk.Entry(frame)
        amount_entry.grid(row=1, column=1, padx=5, pady=5)
        
        tk.Label(frame, text="Date:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=2, column=0, padx=5, pady=5)
        date_entry = ttk.Entry(frame)
        date_entry.grid(row=2, column=1, padx=5, pady=5)
        date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        def update_date():
            top = tk.Toplevel(window)
            cal = Calendar(top, selectmode="day", date_pattern="y-mm-dd")
            cal.pack(pady=10)
            def set_date():
                date_entry.delete(0, tk.END)
                date_entry.insert(0, cal.get_date())
                top.destroy()
            ttk.Button(top, text="Select", command=set_date).pack(pady=5)
        
        ttk.Button(frame, text="Date", command=update_date, width=5).grid(row=2, column=2, padx=5)
        
        def add_income():
            try:
                amount = amount_entry.get()
                if not amount:
                    raise ValueError("Amount cannot be empty")
                date = date_entry.get()
                income = {
                    "source": source_entry.get(),
                    "amount": float(amount),
                    "date": date,
                    "month": date[:7]
                }
                self.all_income.append(income)
                self.update_charts()
                if hasattr(self, "showing_savings_list") and self.showing_savings_list:
                    self.update_savings_table()
                self.save_data()
                self.month_dropdown['values'] = self.get_month_options()
                window.destroy()
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid numerical amount (e.g., 50.00)")
                
        ttk.Button(window, text="Add Income", command=add_income).pack(pady=20)

    def open_expense_window(self):
        window = tk.Toplevel(self.root)
        window.title("Add Expense")
        window.geometry("460x520")
        window.configure(bg=self.colors["surface"])
        
        tk.Label(window, text="New Expense", font=("Arial", 16, "bold"), 
                fg=self.colors["text"], bg=self.colors["surface"]).pack(pady=10)
        
        frame = tk.Frame(window, bg=self.colors["surface"])
        frame.pack(pady=10)
        
        tk.Label(frame, text="Where:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=0, column=0, padx=5, pady=5, sticky="e")
        where_entry = ttk.Entry(frame)
        where_entry.grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(frame, text="Amount:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=1, column=0, padx=5, pady=5, sticky="e")
        amount_entry = ttk.Entry(frame)
        amount_entry.grid(row=1, column=1, padx=5, pady=5)
        
        tk.Label(frame, text="Date:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=2, column=0, padx=5, pady=5, sticky="e")
        date_entry = ttk.Entry(frame)
        date_entry.grid(row=2, column=1, padx=5, pady=5)
        date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        def update_date():
            top = tk.Toplevel(window)
            cal = Calendar(top, selectmode="day", date_pattern="y-mm-dd")
            cal.pack(pady=10)
            def set_date():
                date_entry.delete(0, tk.END)
                date_entry.insert(0, cal.get_date())
                top.destroy()
            ttk.Button(top, text="Select", command=set_date).pack(pady=5)
        
        ttk.Button(frame, text="Date", command=update_date, width=5).grid(row=2, column=2, padx=5)
        
        tk.Label(frame, text="Category:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=3, column=0, padx=5, pady=5, sticky="e")
        category_var = tk.StringVar()
        category_combo = ttk.Combobox(frame, textvariable=category_var, values=self.get_expense_category_options())
        category_combo.grid(row=3, column=1, padx=5, pady=5)
        category_combo.set("Other")
        tk.Label(frame, text="Type a new category to save it for future entries.",
                fg=self.colors["muted"], bg=self.colors["surface"], font=("Segoe UI", 8)).grid(row=4, column=1, columnspan=2, sticky="w", padx=5)

        tk.Label(frame, text="Payment:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=5, column=0, padx=5, pady=5, sticky="e")
        payment_var = tk.StringVar(value="Debit Card")
        payment_combo = ttk.Combobox(frame, textvariable=payment_var, values=self.payment_methods, state="readonly")
        payment_combo.grid(row=5, column=1, padx=5, pady=5)

        tk.Label(frame, text="Authorized User:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=6, column=0, padx=5, pady=5, sticky="e")
        authorized_user_entry = ttk.Entry(frame)
        authorized_user_entry.grid(row=6, column=1, padx=5, pady=5)
        tk.Label(frame, text="Use for credit card spending by another cardholder.",
                fg=self.colors["muted"], bg=self.colors["surface"], font=("Segoe UI", 8)).grid(row=7, column=1, columnspan=2, sticky="w", padx=5)
        
        def add_expense():
            try:
                amount = amount_entry.get()
                if not amount:
                    raise ValueError("Amount cannot be empty")
                date = date_entry.get()
                category = self.add_category(category_var.get())
                expense = {
                    "where": where_entry.get(),
                    "amount": float(amount),
                    "date": date,
                    "category": category,
                    "payment_method": payment_var.get(),
                    "authorized_user": authorized_user_entry.get().strip(),
                    "month": date[:7]  # Extract YYYY-MM
                }
                self.all_expenses.append(expense)
                self.update_charts()
                self.save_data()
                self.month_dropdown['values'] = self.get_month_options()
                window.destroy()
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid numerical amount (e.g., 50.00)")
                
        ttk.Button(window, text="Add Expense", command=add_expense).pack(pady=20)

    def open_saving_window(self):
        window = tk.Toplevel(self.root)
        window.title("Add Saving/Investment")
        window.geometry("440x430")
        window.configure(bg=self.colors["surface"])

        tk.Label(window, text="New Saving / Investment", font=("Arial", 16, "bold"),
                fg=self.colors["text"], bg=self.colors["surface"]).pack(pady=10)

        frame = tk.Frame(window, bg=self.colors["surface"])
        frame.pack(pady=10)

        tk.Label(frame, text="Account:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=0, column=0, padx=5, pady=5, sticky="e")
        account_entry = ttk.Entry(frame)
        account_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(frame, text="Amount:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=1, column=0, padx=5, pady=5, sticky="e")
        amount_entry = ttk.Entry(frame)
        amount_entry.grid(row=1, column=1, padx=5, pady=5)

        tk.Label(frame, text="Date:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=2, column=0, padx=5, pady=5, sticky="e")
        date_entry = ttk.Entry(frame)
        date_entry.grid(row=2, column=1, padx=5, pady=5)
        date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))

        def update_date():
            top = tk.Toplevel(window)
            cal = Calendar(top, selectmode="day", date_pattern="y-mm-dd")
            cal.pack(pady=10)

            def set_date():
                date_entry.delete(0, tk.END)
                date_entry.insert(0, cal.get_date())
                top.destroy()

            ttk.Button(top, text="Select", command=set_date).pack(pady=5)

        ttk.Button(frame, text="Date", command=update_date, width=5).grid(row=2, column=2, padx=5)

        tk.Label(frame, text="Type:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=3, column=0, padx=5, pady=5, sticky="e")
        category_var = tk.StringVar(value="Emergency Fund")
        category_combo = ttk.Combobox(frame, textvariable=category_var, values=self.saving_categories)
        category_combo.grid(row=3, column=1, padx=5, pady=5)

        tk.Label(frame, text="Tracked as money not spent, not as an expense.",
                fg=self.colors["muted"], bg=self.colors["surface"], font=("Segoe UI", 8)).grid(row=4, column=1, columnspan=2, sticky="w", padx=5)

        def add_saving():
            try:
                amount = amount_entry.get()
                if not amount:
                    raise ValueError("Amount cannot be empty")
                date = date_entry.get()
                category = category_var.get().strip() or "Other"
                if category not in self.saving_categories:
                    self.saving_categories.append(category)
                    self.saving_categories.sort(key=str.lower)
                self.all_savings.append({
                    "account": account_entry.get(),
                    "amount": float(amount),
                    "date": date,
                    "category": category,
                    "month": date[:7]
                })
                self.update_charts()
                self.save_data()
                self.month_dropdown['values'] = self.get_month_options()
                window.destroy()
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid numerical amount (e.g., 50.00)")

        ttk.Button(window, text="Add Saving/Investment", command=add_saving).pack(pady=20)

    def open_goals_window(self):
        window = tk.Toplevel(self.root)
        window.title("Goals & Categories")
        window.geometry("720x560")
        window.minsize(560, 420)
        window.configure(bg=self.colors["surface"])
        window.grid_rowconfigure(1, weight=1)
        window.grid_columnconfigure(0, weight=1)

        header = tk.Frame(window, bg=self.colors["surface"], padx=16, pady=12)
        header.grid(row=0, column=0, sticky="ew")
        tk.Label(header, text="Goals & Categories", font=("Segoe UI", 18, "bold"),
                fg=self.colors["text"], bg=self.colors["surface"]).pack(anchor="w")
        tk.Label(
            header,
            text="Add a monthly goal for any spending category as a dollar amount or as a percent of income. Progress bars show whether you are within the goal or over it.",
            font=("Segoe UI", 9),
            fg=self.colors["muted"],
            bg=self.colors["surface"],
            wraplength=560,
            justify="left",
        ).pack(anchor="w", pady=(4, 0))

        category_frame = tk.Frame(window, bg=self.colors["surface"], padx=16)
        category_frame.grid(row=1, column=0, sticky="nsew")
        category_frame.grid_rowconfigure(2, weight=1)
        category_frame.grid_columnconfigure(0, weight=1)
        tk.Label(category_frame, text="Category Goals", font=("Segoe UI", 12, "bold"),
                fg=self.colors["text"], bg=self.colors["surface"]).grid(row=0, column=0, sticky="w", pady=(0, 6))
        tk.Label(
            category_frame,
            text="Select an existing category or type a new one, enter the limit, choose Dollars or Percent of Income, then click Add / Update Goal.",
            font=("Segoe UI", 9),
            fg=self.colors["muted"],
            bg=self.colors["surface"],
            wraplength=560,
            justify="left",
        ).grid(row=1, column=0, sticky="ew", pady=(0, 8))

        tree_frame = tk.Frame(category_frame, bg=self.colors["surface"])
        tree_frame.grid(row=2, column=0, sticky="nsew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree = ttk.Treeview(tree_frame, columns=("Category", "Goal", "Type"), show="headings", height=8)
        tree.heading("Category", text="Category")
        tree.heading("Goal", text="Goal")
        tree.heading("Type", text="Type")
        tree.column("Category", width=240, anchor="w")
        tree.column("Goal", width=130, anchor="e")
        tree.column("Type", width=130, anchor="w")
        tree.grid(row=0, column=0, sticky="nsew")
        tree_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree_scrollbar.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=tree_scrollbar.set)

        def refresh_tree():
            for item in tree.get_children():
                tree.delete(item)
            for category in self.categories:
                goal = self.category_goals.get(category, "")
                goal_type = self.category_goal_types.get(category, "amount")
                if isinstance(goal, float) and goal > 0:
                    goal_text = f"{goal:.2f}%" if goal_type == "percent" else f"${goal:.2f}"
                    goal_type_text = "Percent of Income" if goal_type == "percent" else "Dollars"
                else:
                    goal_text = ""
                    goal_type_text = ""
                tree.insert("", "end", values=(category, goal_text, goal_type_text))
            category_combo["values"] = self.categories

        edit_frame = tk.Frame(window, bg=self.colors["surface"], padx=16, pady=10)
        edit_frame.grid(row=2, column=0, sticky="ew")
        edit_frame.grid_columnconfigure(1, weight=1)
        tk.Label(edit_frame, text="Category:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=0, column=0, padx=4, pady=4)
        category_var = tk.StringVar()
        category_combo = ttk.Combobox(edit_frame, textvariable=category_var, values=self.categories, width=24)
        category_combo.grid(row=0, column=1, padx=4, pady=4, sticky="ew")
        tk.Label(edit_frame, text="Goal:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=0, column=2, padx=4, pady=4)
        goal_entry = ttk.Entry(edit_frame, width=12)
        goal_entry.grid(row=0, column=3, padx=4, pady=4)
        goal_type_var = tk.StringVar(value="Dollars")
        goal_type_combo = ttk.Combobox(
            edit_frame,
            textvariable=goal_type_var,
            values=["Dollars", "Percent of Income"],
            state="readonly",
            width=18
        )
        goal_type_combo.grid(row=0, column=4, padx=4, pady=4)

        def select_category(_event=None):
            selected = tree.selection()
            if not selected:
                return
            category, goal_text, goal_type_text = tree.item(selected[0])["values"]
            category_var.set(category)
            goal_entry.delete(0, tk.END)
            goal_entry.insert(0, str(goal_text).replace("$", "").replace("%", ""))
            goal_type_var.set(goal_type_text or "Dollars")

        tree.bind("<<TreeviewSelect>>", select_category)

        def save_category():
            try:
                category_name = category_var.get().strip()
                if not category_name:
                    messagebox.showwarning("Warning", "Enter or select a category first.")
                    return
                category = self.add_category(category_name)
                goal_text = goal_entry.get().strip()
                if not goal_text:
                    messagebox.showwarning("Warning", "Enter a goal for this category.")
                    return
                goal = float(goal_text)
                if goal <= 0:
                    messagebox.showwarning("Warning", "Goal must be greater than zero.")
                    return
                goal_type = "percent" if goal_type_var.get() == "Percent of Income" else "amount"
                if goal_type == "percent" and goal > 100:
                    messagebox.showwarning("Warning", "Percent of income goals should be 100 or less.")
                    return
                self.category_goals[category] = goal
                self.category_goal_types[category] = goal_type
                self.save_data()
                refresh_tree()
                self.update_charts()
            except ValueError:
                messagebox.showerror("Error", "Please enter goals as numbers, such as 250.00.")

        def delete_category():
            category = category_var.get().strip()
            if not category:
                messagebox.showwarning("Warning", "Select a category to delete.")
                return
            if any(exp["category"] == category for exp in self.all_expenses):
                messagebox.showwarning("In Use", "This category is used by expenses and cannot be deleted.")
                return
            if category in self.categories:
                self.categories.remove(category)
                self.category_goals.pop(category, None)
                self.category_goal_types.pop(category, None)
                self.deleted_categories.add(category)
                self.save_data()
                refresh_tree()
                self.update_charts()
                category_var.set("")
                goal_entry.delete(0, tk.END)

        footer = tk.Frame(window, bg=self.colors["surface"], padx=16, pady=10)
        footer.grid(row=3, column=0, sticky="ew")

        def clear_goal():
            category = category_var.get().strip()
            if not category:
                messagebox.showwarning("Warning", "Select a category first.")
                return
            self.category_goals.pop(category, None)
            self.category_goal_types.pop(category, None)
            self.save_data()
            refresh_tree()
            self.update_charts()
            goal_entry.delete(0, tk.END)

        ttk.Button(footer, text="Add / Update Goal", command=save_category, style="TButton").pack(side="left", padx=(0, 8), pady=2)
        ttk.Button(footer, text="Clear Goal", command=clear_goal, style="TButton").pack(side="left", padx=(0, 8), pady=2)
        ttk.Button(footer, text="Delete Unused Category", command=delete_category, style="TButton").pack(side="left", padx=(0, 8), pady=2)
        ttk.Button(footer, text="Close", command=window.destroy, style="TButton").pack(side="right")
        refresh_tree()

    def update_charts(self):
        income, expenses = self.get_month_data(self.current_month)
        savings = self.get_month_savings(self.current_month)
        total_income = sum(i["amount"] for i in income)
        total_expense = sum(e["amount"] for e in expenses)
        total_savings = self.get_savings_investing_total(expenses, savings)
        net_income = total_income - total_expense

        if hasattr(self, "summary_labels"):
            self.summary_labels["income"].config(text=f"${total_income:.2f}", fg=self.colors["good"])
            self.summary_labels["expenses"].config(text=f"${total_expense:.2f}", fg=self.colors["bad"])
            self.summary_labels["net"].config(
                text=f"${net_income:.2f}",
                fg=self.colors["good"] if net_income >= 0 else self.colors["bad"]
            )
            self.summary_labels["savings"].config(text=f"${total_savings:.2f}", fg=self.colors["accent"])
        
        if self.showing_list:
            self.update_expense_table()
            return
        
        self.income_ax.clear()
        if not income:
            self.income_ax.text(0.5, 0.5, "No Income", ha='center', va='center', 
                              color=self.colors["muted"], fontsize=12)
            self.income_center_text = self.income_ax.text(0, 0, "$0.00", 
                                                        **self.center_text_style())
        else:
            sources = [i["source"] for i in income]
            amounts = [i["amount"] for i in income]
            explode = [0] * len(amounts)
            wedges, texts, autotexts = self.income_ax.pie(
                amounts, labels=sources, autopct='', startangle=90, 
                textprops={'color': self.colors["text"], 'fontsize': 10}, colors=self.chart_colors(len(amounts)),
                explode=explode, labeldistance=1.1
            )
            self.income_wedges = wedges
            self.income_autotexts = autotexts
            self.income_amounts = amounts
            self.income_center_text = self.income_ax.text(0, 0, f"${total_income:.2f}", 
                                                        **self.center_text_style())
        
        self.income_ax.set_title(f"Income Breakdown", color=self.colors["text"], pad=20)
        self.income_canvas.draw()
        
        self.expense_ax.clear()
        category_totals = self.get_expense_breakdown_totals(expenses, savings)
        total_breakdown = sum(category_totals.values())
        if not category_totals:
            self.expense_ax.text(0.5, 0.5, "No Expenses", ha='center', va='center', 
                               color=self.colors["muted"], fontsize=12)
            self.expense_center_text = self.expense_ax.text(0, 0, "$0.00", 
                                                          **self.center_text_style())
        else:
            amounts = list(category_totals.values())
            labels = list(category_totals.keys())
            explode = [0] * len(amounts)
            wedges, texts, autotexts = self.expense_ax.pie(
                amounts, labels=labels, autopct='', startangle=90, 
                textprops={'color': self.colors["text"], 'fontsize': 10}, colors=self.chart_colors(len(amounts)),
                explode=explode, labeldistance=1.1
            )
            self.expense_wedges = wedges
            self.expense_autotexts = autotexts
            self.expense_amounts = amounts
            self.expense_categories = labels
            self.expense_center_text = self.expense_ax.text(0, 0, f"${total_breakdown:.2f}",
                                                          **self.center_text_style())
        
        self.expense_ax.set_title(f"Expense Breakdown", color=self.colors["text"], pad=20)
        self.expense_canvas.draw()
        
        self.update_goal_progress(expenses, savings, total_income)

        color = self.colors["good"] if net_income >= 0 else self.colors["bad"]
        self.net_label.config(text=f"Net Income ({self.current_month}): ${net_income:.2f}", fg=color)

    def update_goal_progress(self, expenses, savings, total_income):
        for widget in self.budget_frame.winfo_children():
            widget.destroy()

        wrapper = tk.Frame(self.budget_frame, bg=self.colors["surface"], padx=16, pady=12,
                           highlightbackground=self.colors["border"], highlightthickness=1)
        wrapper.pack(fill="x", padx=24, pady=6)
        tk.Label(wrapper, text="Goal Progress", font=("Segoe UI", 13, "bold"),
                fg=self.colors["text"], bg=self.colors["surface"]).pack(anchor="w")

        list_container = tk.Frame(wrapper, bg=self.colors["surface"])
        list_container.pack(fill="x", expand=False, pady=(6, 0))
        list_canvas = tk.Canvas(list_container, height=150, bg=self.colors["surface"], highlightthickness=0)
        list_canvas.pack(side="left", fill="x", expand=True)
        scrollbar = ttk.Scrollbar(list_container, orient="vertical", command=list_canvas.yview)
        scrollbar.pack(side="right", fill="y")
        list_canvas.configure(yscrollcommand=scrollbar.set)
        rows_frame = tk.Frame(list_canvas, bg=self.colors["surface"])
        rows_window = list_canvas.create_window((0, 0), window=rows_frame, anchor="nw")

        def configure_rows(_event=None):
            list_canvas.configure(scrollregion=list_canvas.bbox("all"))

        def configure_canvas(event):
            list_canvas.itemconfigure(rows_window, width=event.width)

        rows_frame.bind("<Configure>", configure_rows)
        list_canvas.bind("<Configure>", configure_canvas)

        categories_with_goals = [category for category in self.categories if self.category_goals.get(category, 0) > 0]
        for category in categories_with_goals:
            spent = sum(exp["amount"] for exp in expenses if exp["category"] == category)
            if category == "Savings/Investing":
                spent += sum(item["amount"] for item in savings)
            goal_value = self.category_goals.get(category, 0)
            goal_type = self.category_goal_types.get(category, "amount")
            if goal_type == "percent":
                effective_goal = total_income * (goal_value / 100)
                goal_label = f"{goal_value:.2f}% of income (${effective_goal:.2f})"
            else:
                effective_goal = goal_value
                goal_label = f"${effective_goal:.2f}"
            self.add_progress_row(
                rows_frame,
                category,
                spent,
                effective_goal,
                goal_label,
                is_savings_goal=(category == "Savings/Investing")
            )

        if not categories_with_goals:
            tk.Label(
                rows_frame,
                text="No category goals yet. Open Goals & Categories to add a monthly dollar limit for a category.",
                font=("Segoe UI", 10),
                fg=self.colors["muted"],
                bg=self.colors["surface"],
            ).pack(anchor="w", pady=(8, 0))

    def add_progress_row(self, parent, label, spent, goal, goal_label, is_savings_goal=False):
        row = tk.Frame(parent, bg=self.colors["surface"])
        row.pack(fill="x", pady=4)
        top = tk.Frame(row, bg=self.colors["surface"])
        top.pack(fill="x")
        if goal <= 0:
            status = "Waiting for income"
            is_good = False
        elif is_savings_goal:
            difference = spent - goal
            is_good = difference >= 0
            status = f"Met goal by ${difference:.2f}" if is_good else f"Under goal by ${abs(difference):.2f}"
        else:
            over_by = spent - goal
            is_good = over_by <= 0
            remaining = goal - spent
            status = f"Within goal by ${remaining:.2f}" if is_good else f"Over by ${over_by:.2f}"
        status_color = self.colors["good"] if is_good else self.colors["bad"]
        bar_color = self.colors["good"] if is_good else self.colors["bad"]
        goal_text = f"${spent:.2f} / {goal_label} - {status}"
        tk.Label(top, text=label, font=("Segoe UI", 10, "bold"),
                fg=self.colors["text"], bg=self.colors["surface"]).pack(side="left")
        tk.Label(top, text=goal_text, font=("Segoe UI", 10),
                fg=status_color, bg=self.colors["surface"]).pack(side="right")

        bar = tk.Canvas(row, height=14, bg=self.colors["surface"], highlightthickness=0)
        bar.pack(fill="x", pady=(3, 0))

        def draw_bar(_event=None):
            bar.delete("all")
            width = max(1, bar.winfo_width())
            bar.create_rectangle(0, 0, width, 14, fill=self.colors["surface_alt"], outline="")
            ratio = min(spent / goal, 1.0) if goal else 0
            if ratio > 0:
                bar.create_rectangle(0, 0, width * ratio, 14, fill=bar_color, outline="")
            if goal > 0 and spent > goal and not is_savings_goal:
                bar.create_rectangle(width - 4, 0, width, 14, fill=self.colors["bad"], outline="")

        bar.bind("<Configure>", draw_bar)
        draw_bar()

    def update_expense_table(self):
        _, expenses = self.get_month_data(self.current_month)
        savings = self.get_month_savings(self.current_month)
        for item in self.expense_tree.get_children():
            self.expense_tree.delete(item)
        sorted_expenses = sorted(expenses, key=lambda x: x["date"], reverse=True)
        for i, exp in enumerate(sorted_expenses):
            self.expense_tree.insert("", "end", values=(
                exp["date"],
                exp["where"],
                f"${exp['amount']:.2f}",
                exp["category"],
                exp.get("payment_method", "Debit Card"),
                exp.get("authorized_user", "")
            ), tags=('row',))
            self.expense_tree.tag_configure('row', background=self.colors["surface"])
            if i < len(sorted_expenses) - 1:
                self.expense_tree.insert("", "end", values=("", "", "", "", "", ""), tags=('separator',))
                self.expense_tree.tag_configure('separator', background=self.colors["surface_alt"], font=("Arial", 1))
        
        self.expense_ax.clear()
        category_totals = self.get_expense_breakdown_totals(expenses, savings)
        total_breakdown = sum(category_totals.values())
        if not category_totals:
            self.expense_ax.text(0.5, 0.5, "No Expenses", ha='center', va='center', 
                               color="white", fontsize=12)
            self.expense_center_text = self.expense_ax.text(0, 0, "$0.00", 
                                                          **self.center_text_style())
        else:
            amounts = list(category_totals.values())
            labels = list(category_totals.keys())
            explode = [0] * len(amounts)
            wedges, texts, autotexts = self.expense_ax.pie(
                amounts, labels=labels, autopct='', startangle=90,
                textprops={'color': self.colors["text"], 'fontsize': 10}, colors=self.chart_colors(len(amounts)),
                explode=explode, labeldistance=1.1
            )
            self.expense_wedges = wedges
            self.expense_amounts = amounts
            self.expense_categories = labels
            self.expense_center_text = self.expense_ax.text(0, 0, f"${total_breakdown:.2f}",
                                                          **self.center_text_style())
        
        self.expense_ax.set_title(f"Expense Breakdown ({self.current_month})", color=self.colors["text"], pad=20)
        self.expense_canvas.draw()

    def toggle_income_view(self):
        if not hasattr(self, 'showing_income_list'):
            self.showing_income_list = False
        
        if not self.showing_income_list:
            if hasattr(self, 'showing_savings_list') and self.showing_savings_list:
                self.savings_table_frame.grid_remove()
                self.view_savings_button.config(text="View Saved/Invested List")
                self.showing_savings_list = False
                self.income_canvas.get_tk_widget().grid()
                self.expense_canvas.get_tk_widget().grid()
            self.income_canvas.get_tk_widget().grid_remove()
            self.income_table_frame.grid()
            self.view_income_button.config(text="See Pie Chart")
            self.showing_income_list = True
            self.update_income_table()
        else:
            self.income_table_frame.grid_remove()
            self.income_canvas.get_tk_widget().grid()
            self.view_income_button.config(text="View Income List")
            self.showing_income_list = False
            self.update_charts()

    def toggle_expense_view(self):
        if not hasattr(self, 'showing_expense_list'):
            self.showing_expense_list = False
        
        if not self.showing_expense_list:
            if hasattr(self, 'showing_savings_list') and self.showing_savings_list:
                self.savings_table_frame.grid_remove()
                self.view_savings_button.config(text="View Saved/Invested List")
                self.showing_savings_list = False
                self.income_canvas.get_tk_widget().grid()
                self.expense_canvas.get_tk_widget().grid()
            self.expense_canvas.get_tk_widget().grid_remove()
            self.expense_table_frame.grid()
            self.view_expense_button.config(text="See Pie Chart")
            self.showing_expense_list = True
            self.update_expense_table()
        else:
            self.expense_table_frame.grid_remove()
            self.expense_canvas.get_tk_widget().grid()
            self.view_expense_button.config(text="View Expense List")
            self.showing_expense_list = False
            self.update_charts()

    def toggle_savings_view(self):
        if not hasattr(self, 'showing_savings_list'):
            self.showing_savings_list = False

        if not self.showing_savings_list:
            self.income_table_frame.grid_remove()
            self.expense_table_frame.grid_remove()
            self.income_canvas.get_tk_widget().grid()
            self.expense_canvas.get_tk_widget().grid()
            self.view_income_button.config(text="View Income List")
            self.view_expense_button.config(text="View Expense List")
            self.showing_income_list = False
            self.showing_expense_list = False
            self.savings_table_frame.grid()
            self.savings_table_frame.tkraise()
            self.view_savings_button.config(text="See Pie Charts")
            self.showing_savings_list = True
            self.update_savings_table()
        else:
            self.close_savings_view()

    def close_savings_view(self):
        if hasattr(self, "savings_table_frame"):
            self.savings_table_frame.grid_remove()
        if hasattr(self, "view_savings_button"):
            self.view_savings_button.config(text="View Saved/Invested List")
        self.showing_savings_list = False
        self.update_charts()

    def update_income_table(self):
        income, _ = self.get_month_data(self.current_month)
        for item in self.income_tree.get_children():
            self.income_tree.delete(item)
        sorted_income = sorted(income, key=lambda x: x["date"], reverse=True)
        for i, inc in enumerate(sorted_income):
            self.income_tree.insert("", "end", values=(
                inc["date"],
                inc["source"],
                f"${inc['amount']:.2f}"
            ), tags=('row',))
            self.income_tree.tag_configure('row', background=self.colors["surface"])
            if i < len(sorted_income) - 1:
                self.income_tree.insert("", "end", values=("", "", ""), tags=('separator',))
                self.income_tree.tag_configure('separator', background=self.colors["surface_alt"], font=("Arial", 1))

    def update_expense_table(self):
        _, expenses = self.get_month_data(self.current_month)
        for item in self.expense_tree.get_children():
            self.expense_tree.delete(item)
        sorted_expenses = sorted(expenses, key=lambda x: x["date"], reverse=True)
        for i, exp in enumerate(sorted_expenses):
            self.expense_tree.insert("", "end", values=(
                exp["date"],
                exp["where"],
                f"${exp['amount']:.2f}",
                exp["category"],
                exp.get("payment_method", "Debit Card"),
                exp.get("authorized_user", "")
            ), tags=('row',))
            self.expense_tree.tag_configure('row', background=self.colors["surface"])
            if i < len(sorted_expenses) - 1:
                self.expense_tree.insert("", "end", values=("", "", "", "", "", ""), tags=('separator',))
                self.expense_tree.tag_configure('separator', background=self.colors["surface_alt"], font=("Arial", 1))

    def update_savings_table(self):
        savings = self.get_month_savings(self.current_month)
        for item in self.savings_tree.get_children():
            self.savings_tree.delete(item)
        sorted_savings = sorted(savings, key=lambda x: x["date"], reverse=True)
        for i, saving in enumerate(sorted_savings):
            self.savings_tree.insert("", "end", values=(
                saving["date"],
                saving["account"],
                f"${saving['amount']:.2f}",
                saving.get("category", "Other")
            ), tags=('row',))
            self.savings_tree.tag_configure('row', background=self.colors["surface"])
            if i < len(sorted_savings) - 1:
                self.savings_tree.insert("", "end", values=("", "", "", ""), tags=('separator',))
                self.savings_tree.tag_configure('separator', background=self.colors["surface_alt"], font=("Arial", 1))

    def hide_analysis_view(self):
        self.analysis_frame.grid_remove()
        self.income_table_frame.grid_remove()  # Ensure expense table stays hidden
        self.expense_table_frame.grid_remove()  # Ensure expense table stays hidden

        self.income_canvas.get_tk_widget().grid()
        self.view_list_button.config(text="View List", command=self.toggle_expense_view)
        self.showing_list = False
        self.update_charts()

    def on_income_hover(self, event):
        income, _ = self.get_month_data(self.current_month)
        if not income or self.showing_list:
            return
        
        amounts = [i["amount"] for i in income]
        sources = [i["source"] for i in income]
        explode = [0] * len(amounts)
        hovered = False
        
        for i, wedge in enumerate(self.income_wedges):
            if event.inaxes == self.income_ax and wedge.contains_point([event.x, event.y]):
                explode[i] = 0.1
                hovered_text = f"${self.income_amounts[i]:.2f}"
                hovered = True
                break
        display_labels = [sources[i] if (not hovered or explode[i] > 0) else "" for i in range(len(sources))]
        
        self.income_ax.clear()
        wedges, texts, autotexts = self.income_ax.pie(
            amounts, labels=display_labels, autopct='', startangle=90,
            textprops={'color': self.colors["text"], 'fontsize': 10}, colors=self.chart_colors(len(amounts)),
            explode=explode, labeldistance=1.1
        )
        self.income_wedges = wedges
        self.income_autotexts = autotexts
        self.income_amounts = amounts
        total_income = sum(i["amount"] for i in income)
        center_text = hovered_text if hovered else f"${total_income:.2f}"
        self.income_center_text = self.income_ax.text(
            0, 0, center_text, **self.center_text_style()
        )
        self.income_ax.set_title(f"Income Breakdown", color=self.colors["text"], pad=20)
        self.income_canvas.draw()

    def on_expense_hover(self, event):
        income, expenses = self.get_month_data(self.current_month)
        savings = self.get_month_savings(self.current_month)
        category_totals = self.get_expense_breakdown_totals(expenses, savings)
        if not category_totals:
            return
        
        total_income = sum(i["amount"] for i in income)  # Get total income for percentage calculation
        
        amounts = list(category_totals.values())
        labels = list(category_totals.keys())
        explode = [0] * len(amounts)
        hovered_text = None
        hovered = False

        for i, wedge in enumerate(self.expense_wedges):
            if event.inaxes == self.expense_ax and wedge.contains_point([event.x, event.y]):
                explode[i] = 0.1
                percentage = (amounts[i] / total_income) * 100 if total_income > 0 else 0
                hovered_text = f"${amounts[i]:.2f}\n{percentage:.1f}% of Income"
                hovered = True
                break
        display_labels = [labels[i] if (not hovered or explode[i] > 0) else "" for i in range(len(labels))]
        
        self.expense_ax.clear()
        wedges, texts, autotexts = self.expense_ax.pie(
            amounts, labels=display_labels, autopct='', startangle=90,
            textprops={'color': self.colors["text"], 'fontsize': 10}, colors=self.chart_colors(len(amounts)),
            explode=explode, labeldistance=1.1
        )
        self.expense_wedges = wedges
        self.expense_autotexts = autotexts
        self.expense_amounts = amounts
        self.expense_categories = labels
        total_expense = sum(amounts)
        center_text = hovered_text if hovered else f"${total_expense:.2f}"
        self.expense_center_text = self.expense_ax.text(
            0, 0, center_text, **self.center_text_style()
        )
        self.expense_ax.set_title(f"Expense Breakdown", color=self.colors["text"], pad=20)
        self.expense_canvas.draw()

    def on_income_click(self, event):
        income, _ = self.get_month_data(self.current_month)
        if event.inaxes != self.income_ax or not income:
            return
        for i, wedge in enumerate(self.income_wedges):
            if wedge.contains_point([event.x, event.y]):
                source = income[i]["source"]
                source_income = [inc for inc in income if inc["source"] == source]
                self.show_income_source_popup(source, source_income)
                break

    def show_income_source_popup(self, source, income):
        if hasattr(self, "income_detail_window") and self.income_detail_window.winfo_exists():
            self.income_detail_window.destroy()

        total = sum(inc["amount"] for inc in income)
        window = tk.Toplevel(self.root)
        self.income_detail_window = window
        window.title(f"{source} Income")
        window.configure(bg=self.colors["surface"])
        window.transient(self.root)

        chart_widget = self.income_canvas.get_tk_widget()
        chart_widget.update_idletasks()
        x = chart_widget.winfo_rootx() + chart_widget.winfo_width() - 40
        y = chart_widget.winfo_rooty() + 40
        width = 360
        height = 280
        screen_width = window.winfo_screenwidth()
        if x + width > screen_width:
            x = max(20, chart_widget.winfo_rootx() - width + 40)
        window.geometry(f"{width}x{height}+{x}+{y}")

        header = tk.Frame(window, bg=self.colors["surface"], padx=16, pady=12)
        header.pack(fill="x")
        tk.Label(header, text=source, font=("Segoe UI", 16, "bold"),
                fg=self.colors["text"], bg=self.colors["surface"]).pack(anchor="w")
        tk.Label(header, text=f"Total: ${total:.2f}", font=("Segoe UI", 12, "bold"),
                fg=self.colors["accent"], bg=self.colors["surface"]).pack(anchor="w", pady=(4, 0))

        table_frame = tk.Frame(window, bg=self.colors["surface"], padx=12)
        table_frame.pack(fill="both", expand=True)
        detail_tree = ttk.Treeview(table_frame, columns=("Date", "Amount"), show="headings", height=7)
        detail_tree.heading("Date", text="Date")
        detail_tree.heading("Amount", text="Amount")
        detail_tree.column("Date", width=120, anchor="center")
        detail_tree.column("Amount", width=120, anchor="e")
        detail_tree.pack(fill="both", expand=True)
        for inc in sorted(income, key=lambda item: item["date"], reverse=True):
            detail_tree.insert("", "end", values=(inc["date"], f"${inc['amount']:.2f}"))

        footer = tk.Frame(window, bg=self.colors["surface"], padx=12, pady=10)
        footer.pack(fill="x")
        ttk.Button(footer, text="Close", command=window.destroy, style="TButton").pack(side="right")

    def open_savings_list_window(self, event=None):
        _, expenses = self.get_month_data(self.current_month)
        savings = self.get_month_savings(self.current_month)
        savings_expenses = [
            exp for exp in expenses
            if self.normalize_category(exp.get("category", "")) == "Savings/Investing"
        ]
        total = self.get_savings_investing_total(expenses, savings)

        if hasattr(self, "savings_detail_window") and self.savings_detail_window.winfo_exists():
            self.savings_detail_window.destroy()

        window = tk.Toplevel(self.root)
        self.savings_detail_window = window
        window.title(f"Saved / Invested - {self.current_month}")
        window.configure(bg=self.colors["surface"])
        window.transient(self.root)

        if event is not None:
            x = event.widget.winfo_rootx()
            y = event.widget.winfo_rooty() + event.widget.winfo_height() + 8
        else:
            x = self.root.winfo_rootx() + 60
            y = self.root.winfo_rooty() + 120
        width = 620
        height = 360
        screen_width = window.winfo_screenwidth()
        if x + width > screen_width:
            x = max(20, screen_width - width - 20)
        window.geometry(f"{width}x{height}+{x}+{y}")

        header = tk.Frame(window, bg=self.colors["surface"], padx=16, pady=12)
        header.pack(fill="x")
        tk.Label(
            header,
            text="Saved / Invested",
            font=("Segoe UI", 16, "bold"),
            fg=self.colors["text"],
            bg=self.colors["surface"],
        ).pack(anchor="w")
        tk.Label(
            header,
            text=f"Total: ${total:.2f}",
            font=("Segoe UI", 12, "bold"),
            fg=self.colors["accent"],
            bg=self.colors["surface"],
        ).pack(anchor="w", pady=(4, 0))

        table_frame = tk.Frame(window, bg=self.colors["surface"], padx=12)
        table_frame.pack(fill="both", expand=True)
        detail_tree = ttk.Treeview(
            table_frame,
            columns=("Date", "Type", "Account/Place", "Amount", "Notes"),
            show="headings",
            height=8,
        )
        for heading in ("Date", "Type", "Account/Place", "Amount", "Notes"):
            detail_tree.heading(heading, text=heading)
        detail_tree.column("Date", width=90, anchor="center")
        detail_tree.column("Type", width=110, anchor="w")
        detail_tree.column("Account/Place", width=170, anchor="w")
        detail_tree.column("Amount", width=90, anchor="e")
        detail_tree.column("Notes", width=135, anchor="w")
        detail_tree.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=detail_tree.yview)
        scrollbar.pack(side="right", fill="y")
        detail_tree.configure(yscrollcommand=scrollbar.set)

        rows = []
        for item in savings:
            rows.append((
                item.get("date", ""),
                "Transfer",
                item.get("account", ""),
                item.get("amount", 0),
                item.get("category", ""),
            ))
        for exp in savings_expenses:
            card_user = exp.get("payment_method", "Expense Entry")
            if exp.get("authorized_user", ""):
                card_user = f"{card_user} / {exp['authorized_user']}"
            rows.append((
                exp.get("date", ""),
                "Expense Entry",
                exp.get("where", ""),
                exp.get("amount", 0),
                card_user,
            ))

        for date, row_type, account, amount, notes in sorted(rows, key=lambda item: item[0], reverse=True):
            detail_tree.insert("", "end", values=(date, row_type, account, f"${amount:.2f}", notes))

        if not rows:
            detail_tree.insert("", "end", values=("", "No transfers", "", "$0.00", ""))

        footer = tk.Frame(window, bg=self.colors["surface"], padx=12, pady=10)
        footer.pack(fill="x")
        ttk.Button(footer, text="Close", command=window.destroy, style="TButton").pack(side="right")

    def on_expense_click(self, event):
        _, expenses = self.get_month_data(self.current_month)
        category_totals = self.get_expense_breakdown_totals(expenses)
        if event.inaxes != self.expense_ax or not category_totals:
            return
        for i, wedge in enumerate(self.expense_wedges):
            if wedge.contains_point([event.x, event.y]):
                category = self.expense_categories[i]
                category_expenses = [
                    exp for exp in expenses
                    if self.normalize_category(exp.get("category", "")) == category
                ]
                self.show_expense_category_popup(category, category_expenses)
                break

    def show_expense_category_popup(self, category, expenses, savings=None):
        savings = savings or []
        if hasattr(self, "expense_detail_window") and self.expense_detail_window.winfo_exists():
            self.expense_detail_window.destroy()

        total = sum(exp["amount"] for exp in expenses) + sum(item["amount"] for item in savings)
        window = tk.Toplevel(self.root)
        self.expense_detail_window = window
        window.title(f"{category} Expenses")
        window.configure(bg=self.colors["surface"])
        window.transient(self.root)

        chart_widget = self.expense_canvas.get_tk_widget()
        chart_widget.update_idletasks()
        x = chart_widget.winfo_rootx() + chart_widget.winfo_width() - 40
        y = chart_widget.winfo_rooty() + 40
        width = 430
        height = 330
        screen_width = window.winfo_screenwidth()
        if x + width > screen_width:
            x = max(20, chart_widget.winfo_rootx() - width + 40)
        window.geometry(f"{width}x{height}+{x}+{y}")

        header = tk.Frame(window, bg=self.colors["surface"], padx=16, pady=12)
        header.pack(fill="x")
        tk.Label(
            header,
            text=category,
            font=("Segoe UI", 16, "bold"),
            fg=self.colors["text"],
            bg=self.colors["surface"],
        ).pack(anchor="w")
        tk.Label(
            header,
            text=f"Total: ${total:.2f}",
            font=("Segoe UI", 12, "bold"),
            fg=self.colors["accent"],
            bg=self.colors["surface"],
        ).pack(anchor="w", pady=(4, 0))

        table_frame = tk.Frame(window, bg=self.colors["surface"], padx=12)
        table_frame.pack(fill="both", expand=True)
        detail_tree = ttk.Treeview(
            table_frame,
            columns=("Date", "Place", "Amount", "Card/User"),
            show="headings",
            height=8,
        )
        detail_tree.heading("Date", text="Date")
        detail_tree.heading("Place", text="Place")
        detail_tree.heading("Amount", text="Amount")
        detail_tree.heading("Card/User", text="Card/User")
        detail_tree.column("Date", width=88, anchor="center")
        detail_tree.column("Place", width=120, anchor="w")
        detail_tree.column("Amount", width=78, anchor="e")
        detail_tree.column("Card/User", width=125, anchor="w")
        detail_tree.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=detail_tree.yview)
        scrollbar.pack(side="right", fill="y")
        detail_tree.configure(yscrollcommand=scrollbar.set)

        for exp in sorted(expenses, key=lambda item: item["date"], reverse=True):
            card_user = exp.get("payment_method", "Debit Card")
            if exp.get("authorized_user", ""):
                card_user = f"{card_user} / {exp['authorized_user']}"
            detail_tree.insert(
                "",
                "end",
                values=(exp["date"], exp["where"], f"${exp['amount']:.2f}", card_user),
            )
        for item in sorted(savings, key=lambda entry: entry["date"], reverse=True):
            detail_tree.insert(
                "",
                "end",
                values=(item["date"], item["account"], f"${item['amount']:.2f}", "Saving/Investment"),
            )

        footer = tk.Frame(window, bg=self.colors["surface"], padx=12, pady=10)
        footer.pack(fill="x")
        ttk.Button(footer, text="Close", command=window.destroy, style="TButton").pack(side="right")
    
    def edit_income(self):
        selected = self.income_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an income to edit.")
            return
        
        item = self.income_tree.item(selected[0])
        values = item['values']
        if not values or values[0] == "":
            messagebox.showwarning("Warning", "Please select a valid income entry.")
            return
        
        date, source, amount_str = values
        amount = float(amount_str.replace("$", ""))
        
        for i, inc in enumerate(self.all_income):
            if inc["date"] == date and inc["source"] == source and inc["amount"] == amount:
                income_index = i
                break
        else:
            messagebox.showerror("Error", "Income not found.")
            return

        window = tk.Toplevel(self.root)
        window.title("Edit Income")
        window.geometry("400x350")
        window.configure(bg="#2d2d2d")
        
        tk.Label(window, text="Edit Income", font=("Arial", 16, "bold"), 
                fg="white", bg="#2d2d2d").pack(pady=10)
        
        frame = tk.Frame(window, bg="#2d2d2d")
        frame.pack(pady=10)
        
        tk.Label(frame, text="Source:", fg="white", bg="#2d2d2d").grid(row=0, column=0, padx=5, pady=5)
        source_entry = ttk.Entry(frame)
        source_entry.grid(row=0, column=1, padx=5, pady=5)
        source_entry.insert(0, source)
        
        tk.Label(frame, text="Amount:", fg="white", bg="#2d2d2d").grid(row=1, column=0, padx=5, pady=5)
        amount_entry = ttk.Entry(frame)
        amount_entry.grid(row=1, column=1, padx=5, pady=5)
        amount_entry.insert(0, amount)
        
        tk.Label(frame, text="Date:", fg="white", bg="#2d2d2d").grid(row=2, column=0, padx=5, pady=5)
        date_entry = ttk.Entry(frame)
        date_entry.grid(row=2, column=1, padx=5, pady=5)
        date_entry.insert(0, date)
        
        def update_date():
            top = tk.Toplevel(window)
            cal = Calendar(top, selectmode="day", date_pattern="y-mm-dd")
            cal.pack(pady=10)
            def set_date():
                date_entry.delete(0, tk.END)
                date_entry.insert(0, cal.get_date())
                top.destroy()
            ttk.Button(top, text="Select", command=set_date).pack(pady=5)
        
        ttk.Button(frame, text="Date", command=update_date, width=5).grid(row=2, column=2, padx=5)
        
        def save_edit():
            try:
                new_amount = float(amount_entry.get())
                new_date = date_entry.get()
                self.all_income[income_index] = {
                    "source": source_entry.get(),
                    "amount": new_amount,
                    "date": new_date,
                    "month": new_date[:7]
                }
                self.save_data()
                self.update_income_table()
                self.update_charts()
                window.destroy()
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid numerical amount (e.g., 50.00)")
        
        ttk.Button(window, text="Save Changes", command=save_edit).pack(pady=20)

    def edit_expense(self):
        selected = self.expense_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an expense to edit.")
            return
        
        item = self.expense_tree.item(selected[0])
        values = item['values']
        if not values or values[0] == "":
            messagebox.showwarning("Warning", "Please select a valid expense entry.")
            return
        
        date, where, amount_str, category, payment_method, authorized_user = values
        amount = float(amount_str.replace("$", ""))
        
        for i, exp in enumerate(self.all_expenses):
            if (exp["date"] == date and exp["where"] == where and 
                exp["amount"] == amount and exp["category"] == category and
                exp.get("payment_method", "Debit Card") == payment_method and
                exp.get("authorized_user", "") == authorized_user):
                expense_index = i
                break
        else:
            messagebox.showerror("Error", "Expense not found.")
            return

        window = tk.Toplevel(self.root)
        window.title("Edit Expense")
        window.geometry("460x520")
        window.configure(bg=self.colors["surface"])
        
        tk.Label(window, text="Edit Expense", font=("Arial", 16, "bold"), 
                fg=self.colors["text"], bg=self.colors["surface"]).pack(pady=10)
        
        frame = tk.Frame(window, bg=self.colors["surface"])
        frame.pack(pady=10)
        
        tk.Label(frame, text="Where:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=0, column=0, padx=5, pady=5, sticky="e")
        where_entry = ttk.Entry(frame)
        where_entry.grid(row=0, column=1, padx=5, pady=5)
        where_entry.insert(0, where)
        
        tk.Label(frame, text="Amount:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=1, column=0, padx=5, pady=5, sticky="e")
        amount_entry = ttk.Entry(frame)
        amount_entry.grid(row=1, column=1, padx=5, pady=5)
        amount_entry.insert(0, amount)
        
        tk.Label(frame, text="Date:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=2, column=0, padx=5, pady=5, sticky="e")
        date_entry = ttk.Entry(frame)
        date_entry.grid(row=2, column=1, padx=5, pady=5)
        date_entry.insert(0, date)
        
        def update_date():
            top = tk.Toplevel(window)
            cal = Calendar(top, selectmode="day", date_pattern="y-mm-dd")
            cal.pack(pady=10)
            def set_date():
                date_entry.delete(0, tk.END)
                date_entry.insert(0, cal.get_date())
                top.destroy()
            ttk.Button(top, text="Select", command=set_date).pack(pady=5)
        
        ttk.Button(frame, text="Date", command=update_date, width=5).grid(row=2, column=2, padx=5)
        
        tk.Label(frame, text="Category:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=3, column=0, padx=5, pady=5, sticky="e")
        category_var = tk.StringVar(value=category)
        category_combo = ttk.Combobox(frame, textvariable=category_var, values=self.get_expense_category_options(category))
        category_combo.grid(row=3, column=1, padx=5, pady=5)

        tk.Label(frame, text="Payment:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=4, column=0, padx=5, pady=5, sticky="e")
        payment_var = tk.StringVar(value=payment_method)
        payment_combo = ttk.Combobox(frame, textvariable=payment_var, values=self.payment_methods, state="readonly")
        payment_combo.grid(row=4, column=1, padx=5, pady=5)

        tk.Label(frame, text="Authorized User:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=5, column=0, padx=5, pady=5, sticky="e")
        authorized_user_entry = ttk.Entry(frame)
        authorized_user_entry.grid(row=5, column=1, padx=5, pady=5)
        authorized_user_entry.insert(0, authorized_user)
        tk.Label(frame, text="Use for credit card spending by another cardholder.",
                fg=self.colors["muted"], bg=self.colors["surface"], font=("Segoe UI", 8)).grid(row=6, column=1, columnspan=2, sticky="w", padx=5)
        
        def save_edit():
            try:
                new_amount = float(amount_entry.get())
                new_date = date_entry.get()
                category = self.add_category(category_var.get())
                self.all_expenses[expense_index] = {
                    "where": where_entry.get(),
                    "amount": new_amount,
                    "date": new_date,
                    "category": category,
                    "payment_method": payment_var.get(),
                    "authorized_user": authorized_user_entry.get().strip(),
                    "month": new_date[:7]
                }
                self.save_data()
                self.update_expense_table()
                self.update_charts()
                window.destroy()
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid numerical amount (e.g., 50.00)")
        
        ttk.Button(window, text="Save Changes", command=save_edit).pack(pady=20)

    def edit_saving(self):
        selected = self.savings_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a saved/invested transfer to edit.")
            return

        item = self.savings_tree.item(selected[0])
        values = item['values']
        if not values or values[0] == "":
            messagebox.showwarning("Warning", "Please select a valid saved/invested transfer.")
            return

        date, account, amount_str, category = values
        amount = float(amount_str.replace("$", ""))

        for i, saving in enumerate(self.all_savings):
            if (saving["date"] == date and saving["account"] == account and
                saving["amount"] == amount and saving.get("category", "Other") == category):
                saving_index = i
                break
        else:
            messagebox.showerror("Error", "Saved/invested transfer not found.")
            return

        window = tk.Toplevel(self.root)
        window.title("Edit Saving/Investment")
        window.geometry("440x430")
        window.configure(bg=self.colors["surface"])

        tk.Label(window, text="Edit Saving / Investment", font=("Arial", 16, "bold"),
                fg=self.colors["text"], bg=self.colors["surface"]).pack(pady=10)

        frame = tk.Frame(window, bg=self.colors["surface"])
        frame.pack(pady=10)

        tk.Label(frame, text="Account:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=0, column=0, padx=5, pady=5, sticky="e")
        account_entry = ttk.Entry(frame)
        account_entry.grid(row=0, column=1, padx=5, pady=5)
        account_entry.insert(0, account)

        tk.Label(frame, text="Amount:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=1, column=0, padx=5, pady=5, sticky="e")
        amount_entry = ttk.Entry(frame)
        amount_entry.grid(row=1, column=1, padx=5, pady=5)
        amount_entry.insert(0, amount)

        tk.Label(frame, text="Date:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=2, column=0, padx=5, pady=5, sticky="e")
        date_entry = ttk.Entry(frame)
        date_entry.grid(row=2, column=1, padx=5, pady=5)
        date_entry.insert(0, date)

        def update_date():
            top = tk.Toplevel(window)
            cal = Calendar(top, selectmode="day", date_pattern="y-mm-dd")
            cal.pack(pady=10)

            def set_date():
                date_entry.delete(0, tk.END)
                date_entry.insert(0, cal.get_date())
                top.destroy()

            ttk.Button(top, text="Select", command=set_date).pack(pady=5)

        ttk.Button(frame, text="Date", command=update_date, width=5).grid(row=2, column=2, padx=5)

        tk.Label(frame, text="Type:", fg=self.colors["text"], bg=self.colors["surface"]).grid(row=3, column=0, padx=5, pady=5, sticky="e")
        category_var = tk.StringVar(value=category)
        category_combo = ttk.Combobox(frame, textvariable=category_var, values=self.saving_categories)
        category_combo.grid(row=3, column=1, padx=5, pady=5)

        def save_edit():
            try:
                new_amount = float(amount_entry.get())
                new_date = date_entry.get()
                new_category = category_var.get().strip() or "Other"
                if new_category not in self.saving_categories:
                    self.saving_categories.append(new_category)
                    self.saving_categories.sort(key=str.lower)
                self.all_savings[saving_index] = {
                    "account": account_entry.get(),
                    "amount": new_amount,
                    "date": new_date,
                    "category": new_category,
                    "month": new_date[:7]
                }
                self.save_data()
                self.update_savings_table()
                self.update_charts()
                self.month_dropdown['values'] = self.get_month_options()
                window.destroy()
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid numerical amount (e.g., 50.00)")

        ttk.Button(window, text="Save Changes", command=save_edit).pack(pady=20)

    def delete_income(self):
        selected = self.income_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an income to delete.")
            return
        
        item = self.income_tree.item(selected[0])
        values = item['values']
        if not values or values[0] == "":
            messagebox.showwarning("Warning", "Please select a valid income entry.")
            return
        
        date, source, amount_str = values
        amount = float(amount_str.replace("$", ""))
        
        if not messagebox.askyesno("Confirm", "Are you sure you want to delete this income?"):
            return
        
        for i, inc in enumerate(self.all_income):
            if inc["date"] == date and inc["source"] == source and inc["amount"] == amount:
                del self.all_income[i]
                break
        
        self.save_data()
        self.update_income_table()
        self.update_charts()

    def delete_expense(self):
        selected = self.expense_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an expense to delete.")
            return
        
        item = self.expense_tree.item(selected[0])
        values = item['values']
        if not values or values[0] == "":
            messagebox.showwarning("Warning", "Please select a valid expense entry.")
            return
        
        date, where, amount_str, category, payment_method, authorized_user = values
        amount = float(amount_str.replace("$", ""))
        
        if not messagebox.askyesno("Confirm", "Are you sure you want to delete this expense?"):
            return
        
        for i, exp in enumerate(self.all_expenses):
            if (exp["date"] == date and exp["where"] == where and 
                exp["amount"] == amount and exp["category"] == category and
                exp.get("payment_method", "Debit Card") == payment_method and
                exp.get("authorized_user", "") == authorized_user):
                del self.all_expenses[i]
                break
        
        self.save_data()
        self.update_expense_table()
        self.update_charts()

    def delete_saving(self):
        selected = self.savings_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a saved/invested transfer to delete.")
            return

        item = self.savings_tree.item(selected[0])
        values = item['values']
        if not values or values[0] == "":
            messagebox.showwarning("Warning", "Please select a valid saved/invested transfer.")
            return

        date, account, amount_str, category = values
        amount = float(amount_str.replace("$", ""))

        if not messagebox.askyesno("Confirm", "Are you sure you want to delete this saved/invested transfer?"):
            return

        for i, saving in enumerate(self.all_savings):
            if (saving["date"] == date and saving["account"] == account and
                saving["amount"] == amount and saving.get("category", "Other") == category):
                del self.all_savings[i]
                break

        self.save_data()
        self.update_savings_table()
        self.update_charts()
        self.month_dropdown['values'] = self.get_month_options()

    def analyze_months(self):
        window = tk.Toplevel(self.root)
        window.title("Analyze Months")
        window.geometry("400x400")
        window.configure(bg="#2d2d2d")
        
        tk.Label(window, text="Analyze Spending", font=("Arial", 16, "bold"), 
                fg="white", bg="#2d2d2d").pack(pady=10)
        
        frame = tk.Frame(window, bg="#2d2d2d")
        frame.pack(pady=10)
        
        # Month 1 Dropdown
        tk.Label(frame, text="Month 1:", fg="white", bg="#2d2d2d").grid(row=0, column=0, padx=5, pady=5)
        month1_var = tk.StringVar(value=self.current_month)
        month1_combo = ttk.Combobox(frame, textvariable=month1_var, values=self.get_month_options(), state="readonly")
        month1_combo.grid(row=0, column=1, padx=5, pady=5)
        
        # Month 2 Dropdown
        tk.Label(frame, text="Month 2:", fg="white", bg="#2d2d2d").grid(row=1, column=0, padx=5, pady=5)
        previous_month = (datetime.strptime(self.current_month, "%Y-%m") - timedelta(days=30)).strftime("%Y-%m")
        month2_var = tk.StringVar(value=previous_month if previous_month in self.get_month_options() else self.get_month_options()[1] if len(self.get_month_options()) > 1 else self.current_month)
        month2_combo = ttk.Combobox(frame, textvariable=month2_var, values=self.get_month_options(), state="readonly")
        month2_combo.grid(row=1, column=1, padx=5, pady=5)
        
        # Category Dropdown
        tk.Label(frame, text="Category:", fg="white", bg="#2d2d2d").grid(row=2, column=0, padx=5, pady=5)
        category_var = tk.StringVar(value="All Categories")
        category_options = ["All Categories"] + self.categories
        category_combo = ttk.Combobox(frame, textvariable=category_var, values=category_options, state="readonly")
        category_combo.grid(row=2, column=1, padx=5, pady=5)
        
        def compare_months():
            try:
                month1 = month1_combo.get()
                month2 = month2_combo.get()
                category = category_var.get()
                _, expenses1 = self.get_month_data(month1)
                _, expenses2 = self.get_month_data(month2)
                
                # Hide income pie chart and show analysis table
                self.income_canvas.get_tk_widget().grid_remove()
                self.income_table_frame.grid_remove()  # Ensure income list is hidden too
                self.analysis_table_frame.grid()
                
                # Clear previous analysis
                for item in self.analysis_tree.get_children():
                    self.analysis_tree.delete(item)
                
                if category == "All Categories":
                    # Total spending comparison
                    month1_total = sum(exp["amount"] for exp in expenses1)
                    month2_total = sum(exp["amount"] for exp in expenses2)
                    self.analysis_tree.insert("", "end", text="✨ MONTHLY SPENDING ANALYSIS ✨")
                    self.analysis_tree.insert("", "end", text="═════════════════════════════")
                    self.analysis_tree.insert("", "end", text=f"📅 {month1}: 💰 ${month1_total:.2f}")
                    self.analysis_tree.insert("", "end", text=f"📅 {month2}: 💰 ${month2_total:.2f}")
                    self.analysis_tree.insert("", "end", text="═════════════════════════════")
                    if month1_total > month2_total:
                        self.analysis_tree.insert("", "end", text=f"📈 {month1} outspent {month2} by ${month1_total - month2_total:.2f}")
                    elif month2_total > month1_total:
                        self.analysis_tree.insert("", "end", text=f"📈 {month2} outspent {month1} by ${month2_total - month1_total:.2f}")
                    else:
                        self.analysis_tree.insert("", "end", text="⚖️ Total spending is equal!")
                    self.analysis_tree.insert("", "end", text="═════════════════════════════")
                    
                    # Individual category comparison
                    self.analysis_tree.insert("", "end", text="🔍 Category Breakdown")
                    self.analysis_tree.insert("", "end", text="═════════════════════════════")
                    for cat in self.categories:
                        month1_cat_total = sum(exp["amount"] for exp in expenses1 if exp["category"] == cat)
                        month2_cat_total = sum(exp["amount"] for exp in expenses2 if exp["category"] == cat)
                        if month1_cat_total > 0 or month2_cat_total > 0:
                            self.analysis_tree.insert("", "end", text=f"🏷️ {cat.upper()}")
                            self.analysis_tree.insert("", "end", text=f"  📅 {month1}: 💵 ${month1_cat_total:.2f}")
                            self.analysis_tree.insert("", "end", text=f"  📅 {month2}: 💵 ${month2_cat_total:.2f}")
                            if month1_cat_total > month2_cat_total:
                                diff = month1_cat_total - month2_cat_total
                                self.analysis_tree.insert("", "end", text=f"  📈 {month1} outspent {month2} by ${diff:.2f}")
                            elif month2_cat_total > month1_cat_total:
                                diff = month2_cat_total - month1_cat_total
                                self.analysis_tree.insert("", "end", text=f"  📈 {month2} outspent {month1} by ${diff:.2f}")
                            else:
                                self.analysis_tree.insert("", "end", text="  ⚖️ Even spending")
                            self.analysis_tree.insert("", "end", text="─" * 30)
                else:
                    # Specific category comparison
                    month1_total = sum(exp["amount"] for exp in expenses1 if exp["category"] == category)
                    month2_total = sum(exp["amount"] for exp in expenses2 if exp["category"] == category)
                    self.analysis_tree.insert("", "end", text=f"✨ ANALYSIS FOR {category.upper()} ✨")
                    self.analysis_tree.insert("", "end", text="═════════════════════════════")
                    self.analysis_tree.insert("", "end", text=f"📅 {month1}: 💰 ${month1_total:.2f}")
                    self.analysis_tree.insert("", "end", text=f"📅 {month2}: 💰 ${month2_total:.2f}")
                    self.analysis_tree.insert("", "end", text="═════════════════════════════")
                    if month1_total > month2_total:
                        self.analysis_tree.insert("", "end", text=f"📈 {month1} outspent {month2} by ${month1_total - month2_total:.2f}")
                    elif month2_total > month1_total:
                        self.analysis_tree.insert("", "end", text=f"📈 {month2} outspent {month1} by ${month2_total - month1_total:.2f}")
                    else:
                        self.analysis_tree.insert("", "end", text="⚖️ Even Spending")
                
                window.destroy()
            except Exception as e:
                print(f"Error: {e}")
                messagebox.showerror("Error", "Invalid input. Please check month format (YYYY-MM) and ensure data exists.")
        
        ttk.Button(window, text="Compare", command=compare_months).pack(pady=20)
    
    def close_analysis(self):
        self.analysis_table_frame.grid_remove()
        if hasattr(self, 'showing_income_list') and self.showing_income_list:
            self.income_table_frame.grid()
        else:
            self.income_canvas.get_tk_widget().grid()
    
    def on_closing(self):
        self.save_data()
        plt.close('all')
        self.root.destroy()
        self.root.quit()

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    tracker = BudgetTracker()
    tracker.run()
